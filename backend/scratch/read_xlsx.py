import openpyxl

wb = openpyxl.load_workbook("Rebuild_Accessory_File_Test_File.xlsx", data_only=True)
ws = wb.active

headers = [cell.value for cell in ws[1]]
comp_idx = None
for idx, h in enumerate(headers):
    if h and "compat" in h.lower():
        comp_idx = idx

if comp_idx is not None:
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx == 0:
            continue
        if r_idx < 20:
            print(f"Row {r_idx + 1}: SKU: {row[5]} | Category: {row[4]} | Compatibility: '{row[comp_idx]}'")
        else:
            break
wb.close()
